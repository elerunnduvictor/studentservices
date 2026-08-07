"""
Generate supabase/patch-07-workbook-tabs.sql from the two source workbooks.

The first import took two tabs out of the Org Directory and one out of the KPIs
workbook and left the rest behind. This brings across the tabs that were missed,
keeping each one as its own sheet so the PM Hub matches the workbooks tab for
tab.

Two different situations, handled differently:

  The four department tabs in the Org Directory hold no people the Employee
  Directory does not already have — checked, 105 of 105 present. Copying them
  into their own tables would mean four more places for the same person to go
  stale. They become *filtered sheets* over `employees` instead: same tab, same
  columns, one underlying row.

  The department tabs in the KPIs workbook are the opposite. They are a matrix —
  one row per employee, one column per outcome category — and they carry KPI
  text that never reached the ScoreCard. That is real content, so it gets its
  own table.

Also brought across: the department descriptions, the canonical department and
sub-department lists that drive the dropdowns, and the KPI taxonomy that gives
Speed/Quality/Cost/Autonomy/Satisfaction/Completion their definitions.

    python supabase/build_sheets_patch.py
"""
import pathlib
import sys

import openpyxl

REPO = pathlib.Path(__file__).resolve().parent.parent
DIRECTORY_XLSX = REPO / "Student Services Org Directory.xlsx"
KPIS_XLSX = REPO / "Student Services KPIs.xlsx"
OUT = REPO / "supabase" / "patch-07-workbook-tabs.sql"
OUT8 = REPO / "supabase" / "patch-08-baselines.sql"

# The workbook spells departments its own way; the database settled on these.
CANON = {
    "records, registration & support": "Student Records, Registration, and Support",
    "records registration & support": "Student Records, Registration, and Support",
    "records, registration, support": "Student Records, Registration, and Support",
    "enrollment & retention": "Enrollment & Retention",
    "digital operations": "Digital Operations",
    "digital ops": "Digital Operations",
    "dean of students": "Dean of Students",
}

CATEGORY_COLUMNS = [
    ("speed", "Speed"),
    ("quality", "Quality"),
    ("cost", "Cost"),
    ("student_autonomy", "Student Autonomy"),
    ("student_satisfaction", "Student Satisfaction"),
    ("completion", "Completion"),
]


def norm(v):
    return " ".join(str(v or "").split()).lower()


def sql(v):
    if v is None or v == "":
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"


def grid(ws):
    return [[("" if c is None else str(c).strip()) for c in row]
            for row in ws.iter_rows(values_only=True)]


def cell(row, i):
    return row[i].strip() if i < len(row) and row[i] else ""


def header_row(rows, first_label):
    for i, r in enumerate(rows):
        if any(norm(c) == norm(first_label) for c in r):
            return i
    return None


# ── Org Directory ──────────────────────────────────────────────────────────

def read_departments(wb):
    """Title + description from the top of each department tab."""
    out = []
    tabs = ["Digital Operations", "Dean of Students",
            "Enrollment & Retention", "Records Registration & Support"]
    for order, tab in enumerate(tabs, start=1):
        rows = grid(wb[tab])
        hi = header_row(rows, "NAME")
        title = next((c for c in rows[0] if c), tab) if rows else tab
        desc = ""
        if hi and hi >= 2:
            desc = max((c for c in rows[1] if c), key=len, default="")
        out.append({
            "name": CANON.get(norm(title), title),
            "tab_label": tab,
            "description": desc,
            "sort_order": order,
        })
    return out


def read_sub_departments(wb):
    """The Control tab drives the department / sub-department dropdowns."""
    rows = grid(wb["Control"])
    hi = header_row(rows, "All Department")
    if hi is None:
        sys.exit("Control: could not find the 'All Department' header")
    dcol = next(i for i, c in enumerate(rows[hi]) if norm(c) == "all department")
    pairs, seen = [], set()
    for r in rows[hi + 1:]:
        dept, sub = cell(r, dcol), cell(r, dcol + 1)
        if not dept or not sub:
            continue
        key = (norm(dept), norm(sub))
        if key in seen:
            continue
        seen.add(key)
        pairs.append({"department": CANON.get(norm(dept), dept),
                      "name": sub, "sort_order": len(pairs) + 1})
    return pairs


# ── KPIs workbook ──────────────────────────────────────────────────────────

def read_kpi_categories(wb):
    """
    The Overview tab: six categories under two outcome groups, each with a
    definition and examples. The scorecard groups by these but has never
    carried what they mean.
    """
    rows = grid(wb["Overview"])
    groups, names = rows[0], rows[1]
    defs = next((r for r in rows if norm(r[0]) == "definition"), [])
    exs = next((r for r in rows if norm(r[0]) == "examples"), [])

    out, group = [], ""
    for i, name in enumerate(names):
        if i < len(groups) and groups[i]:
            group = " ".join(groups[i].split())
        if not name or i == 0:
            continue
        out.append({
            "outcome_group": group,
            "name": " ".join(name.split()),
            "definition": cell(defs, i),
            "examples": cell(exs, i),
            "sort_order": len(out) + 1,
        })
    return out


def read_guiding_questions(wb):
    """The questions used to draw KPIs out of each role."""
    rows = grid(wb["Overview"])
    hi = next((i for i, r in enumerate(rows)
               if any(norm(c) == "question" for c in r)), None)
    if hi is None:
        return []
    h = rows[hi]
    qi = next(i for i, c in enumerate(h) if norm(c) == "question")
    gi = next((i for i, c in enumerate(h) if norm(c) == "goal"), None)
    fi = next((i for i, c in enumerate(h) if norm(c) == "focus"), None)
    out = []
    for r in rows[hi + 1:]:
        q = cell(r, qi)
        if not q:
            continue
        out.append({
            "question": q,
            "goal": cell(r, gi) if gi is not None else "",
            "focus": cell(r, fi) if fi is not None else "",
            "sort_order": len(out) + 1,
        })
    return out


def read_baselines(wb):
    """
    The 2025 Baseline KPIs tab: the groundwork behind every KPI — what a role is
    responsible for, what it is working on, what has to be true for it to
    succeed, and where the numbers would come from. Nothing else in the workbook
    records this, and it is the context that makes a bare KPI readable.
    """
    rows = grid(wb["2025 Baseline KPIs"])
    hi = header_row(rows, "Employee Name")
    if hi is None:
        sys.exit("2025 Baseline KPIs: no 'Employee Name' header")
    h = rows[hi]
    want = [
        ("employee_name", "Employee Name"), ("role", "Role"), ("department", "Department"),
        ("scheduling", "Scheduling"), ("key_responsibilities", "Key Responsibilities"),
        ("current_projects", "Current Projects (If Any)"),
        ("conditions_for_success", "Conditions for Success"),
        ("current_kpis", "Current KPIs"), ("suggested_kpis", "Suggested KPIs"),
        ("data_source", "Data Source"), ("timeframe", "Timeframe"),
    ]
    idx = {}
    for key, label in want:
        found = [i for i, c in enumerate(h) if norm(c) == norm(label)]
        if not found:  # tolerate a reworded header rather than losing the column
            found = [i for i, c in enumerate(h) if norm(label).split("(")[0].strip() in norm(c)]
        idx[key] = found[0] if found else None

    out = []
    for r in rows[hi + 1:]:
        name = cell(r, idx["employee_name"])
        if not name:
            continue
        rec = {"sort_order": len(out) + 1}
        for key in idx:
            rec[key] = cell(r, idx[key]) if idx[key] is not None else ""
        rec["department"] = CANON.get(norm(rec.get("department")), rec.get("department"))
        out.append(rec)
    return out


def read_registry_extras(wb):
    """
    KPIs the Registry lists that never reached the ScoreCard.

    The ScoreCard is the newer document — it fills in placeholders the Registry
    left open — so these are not a rollback of it: they are measures identified
    for a role and not yet put on the scorecard. They load as `Not Tracking`,
    which keeps them on the record without touching the hub, since the hub only
    ever shows Tracking rows.
    """
    sc = grid(wb["KPI ScoreCard"])
    h = header_row(sc, "Employee")
    seen = {(norm(r[0]), norm(r[3])) for r in sc[h + 1:]
            if r and r[0] and len(r) > 3 and r[3]}

    reg = grid(wb["KPI Registry"])
    hr = header_row(reg, "Employee")
    cols = ["employee", "role", "department", "kpi_measure", "kpi_category",
            "category_type", "data_availability", "band_green", "band_yellow", "band_red"]
    out = []
    for r in reg[hr + 1:]:
        if not r or not r[0] or len(r) < 4 or not r[3]:
            continue
        if (norm(r[0]), norm(r[3])) in seen:
            continue
        rec = {c: cell(r, i) for i, c in enumerate(cols)}
        rec["department"] = CANON.get(norm(rec["department"]), rec["department"])
        rec["tracking_status"] = "Not Tracking"
        out.append(rec)
    return out


def flatten_matrix_kpis(matrix, wb):
    """
    Turn the department matrix back into individual KPI rows.

    Each category cell holds one measure per line. Those lines are real KPIs
    that were written down for a role and never made it onto the ScoreCard —
    Brad Lester's case response rates, Kira Hayes' QA and KB team costs, and
    others. They belong in `kpis` so they can be picked up and tracked when the
    time comes, and they load as `Not Tracking` so nothing reaches the hub until
    somebody decides it should.
    """
    sc = grid(wb["KPI ScoreCard"])
    h = header_row(sc, "Employee")
    known = {(norm(r[0]), norm(r[3])) for r in sc[h + 1:]
             if r and r[0] and len(r) > 3 and r[3]}

    groups = {"speed": "Operational Outcomes", "quality": "Operational Outcomes",
              "cost": "Operational Outcomes", "student_autonomy": "Student Outcomes",
              "student_satisfaction": "Student Outcomes", "completion": "Student Outcomes"}

    out, seen = [], set()
    for row in matrix:
        for key, label in CATEGORY_COLUMNS:
            raw = row.get(key) or ""
            for line in str(raw).split("\n"):
                measure = line.strip().lstrip(">").strip()
                if not measure:
                    continue
                pair = (norm(row["employee_name"]), norm(measure))
                if pair in known or pair in seen:
                    continue
                seen.add(pair)
                out.append({
                    "employee": row["employee_name"],
                    "role": row.get("role") or None,
                    "department": row["department"],
                    "kpi_measure": measure,
                    "kpi_category": groups[key],
                    "category_type": label,
                    "data_availability": None,
                    "tracking_status": "Not Tracking",
                })
    return out


def read_kpi_matrix(wb):
    """
    One row per employee per department, with the six category columns kept as
    they are written in the workbook — multi-line text and all.
    """
    tabs = [
        ("Dean of Students", "Dean of Students"),
        ("Digital Ops", "Digital Operations"),
        ("Enrollment & Retention", "Enrollment & Retention"),
        ("Records, Registration, Support ", "Student Records, Registration, and Support"),
    ]
    out = []
    for tab, dept in tabs:
        if tab not in wb.sheetnames:
            sys.exit(f"KPIs workbook has no tab named {tab!r}")
        rows = grid(wb[tab])
        hi = header_row(rows, "Employee Name")
        if hi is None:
            sys.exit(f"{tab}: no 'Employee Name' header")
        h = rows[hi]
        idx = {}
        for key, label in ([("employee_name", "Employee Name"), ("role", "Role"),
                            ("employment_status", "Employment Status"),
                            ("stewardship", "Stewardship")] + CATEGORY_COLUMNS):
            found = [i for i, c in enumerate(h) if norm(c) == norm(label)]
            idx[key] = found[0] if found else None

        n = 0
        for r in rows[hi + 1:]:
            name = cell(r, idx["employee_name"])
            if not name or norm(name) in ("0", "total"):
                continue
            n += 1
            rec = {"department": dept, "tab_label": tab.strip(), "sort_order": n}
            for key in idx:
                rec[key] = cell(r, idx[key]) if idx[key] is not None else ""
            out.append(rec)
    return out


# ── emit ───────────────────────────────────────────────────────────────────

def insert(table, cols, records, conflict=None):
    if not records:
        return f"-- {table}: nothing to load\n\n"
    lines = [f"insert into public.{table} ({', '.join(cols)}) values"]
    vals = [",\n".join("  (" + ", ".join(sql(r.get(c)) for c in cols) + ")" for r in records)]
    lines.append(vals[0])
    if conflict:
        upd = ", ".join(f"{c} = excluded.{c}" for c in cols if c not in conflict)
        lines.append(f"on conflict ({', '.join(conflict)}) do update set {upd}")
    return "\n".join(lines) + ";\n\n"


def main():
    for p in (DIRECTORY_XLSX, KPIS_XLSX):
        if not p.exists():
            sys.exit(f"Missing workbook: {p.name}\nPut it in the repository root and re-run.")

    dwb = openpyxl.load_workbook(DIRECTORY_XLSX, data_only=True, read_only=True)
    kwb = openpyxl.load_workbook(KPIS_XLSX, data_only=True, read_only=True)

    departments = read_departments(dwb)
    subs = read_sub_departments(dwb)
    categories = read_kpi_categories(kwb)
    questions = read_guiding_questions(kwb)
    matrix = read_kpi_matrix(kwb)

    L = ['''-- ═══════════ PATCH 07 — THE TABS THE FIRST IMPORT LEFT BEHIND ═══════════
--
-- The Org Directory workbook has eight tabs and the KPIs workbook twelve. The
-- first import took two from one and one from the other, which is why the PM
-- Hub looked nothing like the workbooks it replaced.
--
-- What arrives here, and why each is shaped the way it is:
--
--   departments          the title and description at the top of each
--                        department tab — the only content on those tabs that
--                        is not already in the Employee Directory. Checked
--                        before assuming: all 105 people listed across the four
--                        tabs are in the directory, so the tabs themselves
--                        become filtered views of `employees` rather than four
--                        more copies of the same roster that could drift apart.
--
--   sub_departments      the Control tab. These lists were hardcoded in the PM
--                        Hub's JavaScript; in the database an edit to them
--                        reaches the dropdowns without a deploy.
--
--   kpi_categories       the Overview tab. Speed, Quality, Cost, Autonomy,
--                        Satisfaction and Completion with their definitions and
--                        examples. The scorecard has always grouped by these
--                        without recording what they mean.
--
--   kpi_guiding_questions  the questions used to draw KPIs out of a role.
--
--   department_kpi_matrix  the per-department KPI tabs. Unlike the directory
--                        tabs these are *not* a view of anything: they are a
--                        matrix of employee against outcome category, and they
--                        contain KPI text that never reached the ScoreCard. So
--                        they get their own table and stay editable as written.
--
-- Run after patch-06. Safe to re-run.

begin;

-- ── departments ───────────────────────────────────────────────────────────
create table if not exists public.departments (
  name        text primary key,
  tab_label   text,                    -- what the workbook tab is called
  description text,
  sort_order  integer not null default 0,
  updated_at  timestamptz not null default now(),
  updated_by  text
);

create table if not exists public.sub_departments (
  id          bigserial primary key,
  department  text not null,
  name        text not null,
  sort_order  integer not null default 0,
  updated_at  timestamptz not null default now(),
  updated_by  text,
  unique (department, name)
);

-- ── KPI reference ─────────────────────────────────────────────────────────
create table if not exists public.kpi_categories (
  name          text primary key,
  outcome_group text,                  -- Operational Outcomes | Student Outcomes
  definition    text,
  examples      text,
  sort_order    integer not null default 0,
  updated_at    timestamptz not null default now(),
  updated_by    text
);

create table if not exists public.kpi_guiding_questions (
  id         bigserial primary key,
  question   text not null,
  goal       text,
  focus      text,
  sort_order integer not null default 0,
  updated_at timestamptz not null default now(),
  updated_by text
);

-- ── the per-department KPI matrix ─────────────────────────────────────────
-- Six category columns held as written, multi-line text and all: this is a
-- working document, and reformatting it would lose the author's structure.
create table if not exists public.department_kpi_matrix (
  id                   bigserial primary key,
  department           text not null,
  tab_label            text,
  sort_order           integer not null default 0,
  employee_name        text not null,
  role                 text,
  employment_status    text,
  stewardship          text,
  speed                text,
  quality              text,
  cost                 text,
  student_autonomy     text,
  student_satisfaction text,
  completion           text,
  created_at           timestamptz not null default now(),
  updated_at           timestamptz not null default now(),
  updated_by           text
);

''']

    # plumbing
    L.append("""-- ── plumbing: same touch/audit/RLS treatment as every other table ─────────
do $$
declare
  t     text;
  nm    text;
  names text[];
  chk constant text :=
    'lower(coalesce((select auth.jwt()) ->> ''email'', '''')) in '
    '(select lower(e.email) from public.allowed_editors e)';
begin
  foreach t in array array['departments', 'sub_departments', 'kpi_categories',
                           'kpi_guiding_questions', 'department_kpi_matrix']
  loop
    execute format('drop trigger if exists %1$s_touch on public.%1$s;
      create trigger %1$s_touch before update on public.%1$s
        for each row execute function public.touch_row();', t);

    execute format('drop trigger if exists %1$s_audit on public.%1$s;
      create trigger %1$s_audit after insert or update or delete on public.%1$s
        for each row execute function public.record_change();', t);

    select coalesce(array_agg(policyname), '{}') into names
      from pg_policies where schemaname = 'public' and tablename = t;
    foreach nm in array names loop
      execute format('drop policy if exists %I on public.%I', nm, t);
    end loop;

    execute format('alter table public.%I enable row level security;', t);

    execute format('create policy "%1$s_select" on public.%1$s
        for select using (true);', t);
    execute format('create policy "%1$s_insert" on public.%1$s
        for insert to authenticated with check (%2$s);', t, chk);
    execute format('create policy "%1$s_update" on public.%1$s
        for update to authenticated using (%2$s) with check (%2$s);', t, chk);
    execute format('create policy "%1$s_delete" on public.%1$s
        for delete to authenticated using (%2$s);', t, chk);

    execute format('grant select on public.%I to anon, authenticated;', t);
  end loop;
end;
$$;

-- ── load ──────────────────────────────────────────────────────────────────
truncate public.department_kpi_matrix restart identity;
truncate public.kpi_guiding_questions restart identity;
truncate public.sub_departments restart identity;

""")

    L.append(insert("departments",
                    ["name", "tab_label", "description", "sort_order"],
                    departments, conflict=["name"]))
    L.append(insert("sub_departments", ["department", "name", "sort_order"], subs))
    L.append(insert("kpi_categories",
                    ["name", "outcome_group", "definition", "examples", "sort_order"],
                    categories, conflict=["name"]))
    L.append(insert("kpi_guiding_questions",
                    ["question", "goal", "focus", "sort_order"], questions))
    L.append(insert("department_kpi_matrix",
                    ["department", "tab_label", "sort_order", "employee_name", "role",
                     "employment_status", "stewardship", "speed", "quality", "cost",
                     "student_autonomy", "student_satisfaction", "completion"],
                    matrix))

    L.append('''-- ── what the hub reads ────────────────────────────────────────────────────
create or replace view public.v_hub_departments as
  select d.name, d.tab_label, d.description, d.sort_order,
         (select count(*) from public.employees e where e.department = d.name) as staff_count,
         coalesce((select c.headcount from public.student_contractor_counts c
                    where c.department = d.name), 0) as contractor_count
    from public.departments d
   order by d.sort_order, d.name;

alter view public.v_hub_departments set (security_invoker = on);
grant select on public.v_hub_departments to anon, authenticated;

commit;
''')

    OUT.write_text("".join(L), encoding="utf-8", newline="\n")

    # ── patch 08 ──────────────────────────────────────────────────────────
    baselines = read_baselines(kwb)
    extras = read_registry_extras(kwb)

    B = ['''-- ═══════════ PATCH 08 — BASELINES AND THE REGISTRY REMAINDER ═══════════
--
-- Two things left after patch-07, both decided on evidence rather than taste.
--
-- 1. `kpi_baselines` — the 2025 Baseline KPIs tab.
--    The groundwork behind every KPI: what a role is responsible for, what it
--    is working on, what has to be true for it to succeed, where the numbers
--    would come from and how often they land. Nothing else in either workbook
--    records this, and it is the context that makes a one-line KPI readable.
--
-- 2. The KPI Registry rows that never reached the ScoreCard.
--    Worth being precise about the direction of travel here, because it decides
--    whether adding them back is a restoration or a regression. The ScoreCard
--    is the *newer* document: where the Registry left a measure open, the
--    ScoreCard fills it in — Ana De Castro's endorsement KPI is "X%" with no
--    bands in the older copy and "90%" with full bands on the ScoreCard. So
--    these rows are not something the ScoreCard removed; they are measures
--    identified for a role and not yet placed on it.
--
--    They load as `Not Tracking`, which puts them on the record without
--    touching the hub — the KPI Scorecard page only ever renders Tracking rows.
--    Anything already present is left exactly as it is.
--
-- Run after patch-07. Safe to re-run.

begin;

create table if not exists public.kpi_baselines (
  id                     bigserial primary key,
  sort_order             integer not null default 0,
  employee_name          text not null,
  role                   text,
  department             text,
  scheduling             text,
  key_responsibilities   text,
  current_projects       text,
  conditions_for_success text,
  current_kpis           text,
  suggested_kpis         text,
  data_source            text,
  timeframe              text,
  created_at             timestamptz not null default now(),
  updated_at             timestamptz not null default now(),
  updated_by             text
);

''']

    B.append("""-- same touch/audit/RLS treatment as every other table
do $$
declare
  nm    text;
  names text[];
  chk constant text :=
    'lower(coalesce((select auth.jwt()) ->> ''email'', '''')) in '
    '(select lower(e.email) from public.allowed_editors e)';
begin
  execute 'drop trigger if exists kpi_baselines_touch on public.kpi_baselines;
    create trigger kpi_baselines_touch before update on public.kpi_baselines
      for each row execute function public.touch_row();';
  execute 'drop trigger if exists kpi_baselines_audit on public.kpi_baselines;
    create trigger kpi_baselines_audit after insert or update or delete on public.kpi_baselines
      for each row execute function public.record_change();';

  select coalesce(array_agg(policyname), '{}') into names
    from pg_policies where schemaname = 'public' and tablename = 'kpi_baselines';
  foreach nm in array names loop
    execute format('drop policy if exists %I on public.kpi_baselines', nm);
  end loop;

  execute 'alter table public.kpi_baselines enable row level security';
  execute 'create policy "kpi_baselines_select" on public.kpi_baselines for select using (true)';
  execute format('create policy "kpi_baselines_insert" on public.kpi_baselines
      for insert to authenticated with check (%s)', chk);
  execute format('create policy "kpi_baselines_update" on public.kpi_baselines
      for update to authenticated using (%1$s) with check (%1$s)', chk);
  execute format('create policy "kpi_baselines_delete" on public.kpi_baselines
      for delete to authenticated using (%s)', chk);
  execute 'grant select on public.kpi_baselines to anon, authenticated';
end;
$$;

truncate public.kpi_baselines restart identity;

""")

    B.append(insert("kpi_baselines",
                    ["sort_order", "employee_name", "role", "department", "scheduling",
                     "key_responsibilities", "current_projects", "conditions_for_success",
                     "current_kpis", "suggested_kpis", "data_source", "timeframe"],
                    baselines))

    B.append("""-- ── the Registry remainder ────────────────────────────────────────────────
-- Guarded on employee + measure so a re-run adds nothing twice, and so a row a
-- PM has since edited on the ScoreCard is never overwritten by the older text.
""")
    kcols = ["employee", "role", "department", "kpi_measure", "kpi_category",
             "category_type", "data_availability", "band_green", "band_yellow",
             "band_red", "tracking_status"]
    for r in extras:
        vals = ", ".join(sql(r.get(c)) for c in kcols)
        B.append(
            f"insert into public.kpis ({', '.join(kcols)})\n"
            f"select {vals}\n"
            f" where not exists (select 1 from public.kpis k\n"
            f"   where lower(k.employee) = lower({sql(r['employee'])})\n"
            f"     and lower(k.kpi_measure) = lower({sql(r['kpi_measure'])}));\n"
        )

    B.append("\ncommit;\n")
    OUT8.write_text("".join(B), encoding="utf-8", newline="\n")

    # ── patch 10 ──────────────────────────────────────────────────────────
    promoted = flatten_matrix_kpis(matrix, kwb)
    P = ['''-- ═══════════ PATCH 10 — MATRIX KPIs ONTO THE SCORECARD SHEET ═══════════
--
-- The per-department KPI tabs hold measures that were written down for a role
-- and never reached the ScoreCard: Brad Lester's case response rates, Kira
-- Hayes' QA and Knowledge Base team costs, Matthew Smith's case QA, and more.
-- Until now they existed only inside a matrix cell, which is no use to anyone
-- who wants to start tracking one.
--
-- Each line of each category cell becomes a KPI row, carrying the category the
-- column it came from implies — Speed, Quality and Cost are Operational
-- Outcomes; Autonomy, Satisfaction and Completion are Student Outcomes.
--
-- They load as `Not Tracking`, which is the whole point of the exercise: the PM
-- Hub's KPI ScoreCard sheet reads the table unfiltered and shows all of them,
-- ready to be given bands and switched on, while the hub's scorecard reads
-- v_hub_kpis and continues to show only Tracking rows. Nothing appears in front
-- of the organisation until somebody decides it should.
--
-- Guarded on employee + measure, so re-running adds nothing twice and a row
-- already edited on the ScoreCard is never overwritten.
--
-- Run after patch-09. Safe to re-run.

begin;

''']
    kcols = ["employee", "role", "department", "kpi_measure", "kpi_category",
             "category_type", "data_availability", "tracking_status"]
    for r in promoted:
        vals = ", ".join(sql(r.get(c)) for c in kcols)
        P.append(
            f"insert into public.kpis ({', '.join(kcols)})\n"
            f"select {vals}\n"
            f" where not exists (select 1 from public.kpis k\n"
            f"   where lower(k.employee) = lower({sql(r['employee'])})\n"
            f"     and lower(k.kpi_measure) = lower({sql(r['kpi_measure'])}));\n"
        )
    P.append('''
commit;

-- ── check ─────────────────────────────────────────────────────────────────
-- Tracking must be unchanged; only the Not Tracking pool grows.
select tracking_status, count(*)
  from public.kpis group by tracking_status order by 1;
select count(*) as hub_scorecard_rows from public.v_hub_kpis;
''')
    OUT10 = REPO / "supabase" / "patch-10-matrix-kpis.sql"
    OUT10.write_text("".join(P), encoding="utf-8", newline="\n")

    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  departments            {len(departments)}")
    print(f"  sub_departments        {len(subs)}")
    print(f"  kpi_categories         {len(categories)}")
    print(f"  kpi_guiding_questions  {len(questions)}")
    print(f"  department_kpi_matrix  {len(matrix)}")
    print(f"wrote {OUT8.relative_to(REPO)}")
    print(f"  kpi_baselines          {len(baselines)}")
    print(f"  registry-only KPIs     {len(extras)}  (loaded as Not Tracking)")
    print(f"wrote {OUT10.relative_to(REPO)}")
    print(f"  matrix-only KPIs       {len(promoted)}  (loaded as Not Tracking)")


if __name__ == "__main__":
    main()
