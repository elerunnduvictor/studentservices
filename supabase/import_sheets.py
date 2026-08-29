"""
Read the three source workbooks and emit supabase/seed.sql.

    python supabase/import_sheets.py

Run this once to load Supabase, and again any time the sheets are the source of
truth ahead of the database (it truncates and reloads). After the PM app goes
live the database is the source of truth and this script is only a fallback.

The mapping between each sheet and its table is stated in schema.sql; this file
only deals with locating the data inside sheets that carry banners, legends and
side blocks above and beside the real rows.
"""
import datetime
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "supabase" / "seed.sql"

# Where a workbook might be sitting.
#
# This used to be the single path REPO/"data-sources", a directory that does not
# exist in this repository — so every workbook resolved to a missing file and the
# script could not run at all. Sheets are dropped in the repo root in practice,
# so both are searched, root first.
SEARCH_DIRS = [REPO, REPO / "data-sources"]


def find_workbook(filename):
    """The first place this workbook actually exists, or None."""
    for d in SEARCH_DIRS:
        p = d / filename
        if p.exists():
            return p
    return None


OKR_XLSX = find_workbook("Profit.co Monthly Reports.xlsx")
DIR_XLSX = find_workbook("Student Services Org Directory.xlsx")
KPI_XLSX = find_workbook("Student Services KPIs.xlsx")

PM_EDITORS_JS = REPO / "pm" / "js" / "pm-editors.js"


# ── helpers ────────────────────────────────────────────────────────────────
def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, str):
        v = v.replace("\r\n", "\n").strip()
        return v or None
    return v


def sql(v):
    """Quote a Python value as a SQL literal."""
    if v is None or v == "":
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"


def num(v):
    """Numeric or null — blanks and stray text become null."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return float(m.group(0)) if m else None


def date_or_null(v):
    v = clean(v)
    if not v:
        return None
    return v if re.match(r"^\d{4}-\d{2}-\d{2}$", str(v)) else None


def rows_of(ws, min_row, min_col, max_col):
    for r in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, values_only=True):
        if any(v is not None and str(v).strip() for v in r):
            yield [clean(v) for v in r]


def insert(table, columns, records):
    """Build one multi-row INSERT. Returns '' when there is nothing to write."""
    if not records:
        return ""
    lines = [f"insert into public.{table} ({', '.join(columns)}) values"]
    body = [
        "  (" + ", ".join(sql(v) for v in rec) + ")"
        for rec in records
    ]
    lines.append(",\n".join(body) + ";")
    return "\n".join(lines) + "\n\n"


# ── 1. OKRs ────────────────────────────────────────────────────────────────
def load_okrs():
    ws = openpyxl.load_workbook(OKR_XLSX, data_only=True)["Sheet1"]
    out, order = [], 0
    for r in rows_of(ws, 2, 1, 16):
        if not r[0] and not r[2]:
            continue
        order += 1
        out.append([
            order, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8],
            num(r[9]), num(r[10]), num(r[11]), r[12], r[13], r[14], date_or_null(r[15]),
        ])
    return out


# ── 2. Directory ───────────────────────────────────────────────────────────
def find_header(ws, needle, max_scan=15):
    """Locate the row holding a known column heading."""
    for ri in range(1, min(max_scan, ws.max_row) + 1):
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(ri, ci).value
            if v and str(v).strip().upper() == needle:
                return ri, ci
    raise SystemExit(f'Could not find heading "{needle}" in {ws.title}')


def load_employees():
    ws = openpyxl.load_workbook(DIR_XLSX, data_only=True)["Employee Directory"]
    hdr_row, name_col = find_header(ws, "NAME")
    out, order = [], 0
    # NAME, ROLE, DEPARTMENT, EMPLOYMENT TYPE, PRIMARY STAKEHOLDER,
    # SUB DEPT, CONTRACT ORGANIZATION, Tier Type
    for r in rows_of(ws, hdr_row + 1, name_col, name_col + 7):
        if not r[0]:
            continue
        order += 1
        tier = r[7]
        out.append([
            order, r[0], r[1], r[2], r[3], r[4], r[5], r[6],
            None if tier is None else str(tier),
        ])
    return out


def load_student_employees():
    ws = openpyxl.load_workbook(DIR_XLSX, data_only=True)["Employee Directory"]
    hdr_row, name_col = find_header(ws, "EMPLOYEE NAME")
    out, order = [], 0
    # EMPLOYEE NAME, JOB NAME, Role Title, Sub-Dept, Supervisor, Dept
    for r in rows_of(ws, hdr_row + 1, name_col, name_col + 5):
        if not r[0]:
            continue
        order += 1
        out.append([order, r[0], r[1], r[2], r[3], r[4], r[5]])
    return out


def load_org_chart():
    ws = openpyxl.load_workbook(DIR_XLSX, data_only=True)["Org Chart Info"]
    out, order = [], 0
    for r in rows_of(ws, 3, 1, 10):
        if not r[0]:
            continue
        order += 1
        out.append([order, r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9]])
    return out


# ── 3. KPIs ────────────────────────────────────────────────────────────────
def load_kpis():
    ws = openpyxl.load_workbook(KPI_XLSX, data_only=True)["KPI ScoreCard"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]

    def col(label, default=None):
        for i, h in enumerate(headers):
            if h == label.lower():
                return i
        if default is not None:
            return default
        raise SystemExit(f'Column "{label}" missing from KPI ScoreCard')

    perf = col("performance")
    idx = {
        "employee": col("employee"), "role": col("role"), "dept": col("dept"),
        "measure": col("kpi measure"), "category": col("kpi category"),
        "type": col("category type"), "data": col("data"),
        "green": perf, "yellow": perf + 1, "red": perf + 2,
        "tracking": col("tracking status"), "value": col("current value"),
        "source": col("data source (link to sheet or report)"),
        "freq": col("update frequency"), "update_date": col("update date"),
        "direction": col("direction"), "gcut": col("green cutoff"), "rcut": col("red cutoff"),
    }

    out, order = [], 0
    for r in rows_of(ws, 3, 1, ws.max_column):
        get = lambda k: r[idx[k]] if idx[k] < len(r) else None
        if not get("measure"):
            continue
        order += 1
        tracking = (get("tracking") or "Not Tracking")
        value = get("value")
        out.append([
            order, get("employee"), get("role"), get("dept"), get("measure"),
            get("category"), get("type"), get("data"),
            None if get("green") is None else str(get("green")),
            None if get("yellow") is None else str(get("yellow")),
            None if get("red") is None else str(get("red")),
            str(tracking).strip(),
            None if value is None else str(value),
            get("source"), get("freq"), date_or_null(get("update_date")),
            get("direction"),
            None if get("gcut") is None else str(get("gcut")),
            None if get("rcut") is None else str(get("rcut")),
        ])
    return out


# ── editors ────────────────────────────────────────────────────────────────
def load_editors():
    """
    Who may write, read from pm/js/pm-editors.js — the one list a person edits.

    That file is also what the console's sign-in screen checks, so the message a
    blocked user sees and the rule the database enforces can never drift apart.
    Any trailing // comment on a line is kept as the note, which is where each
    editor's role is recorded.
    """
    if not PM_EDITORS_JS.exists():
        sys.exit(f"Missing {PM_EDITORS_JS} — the console's editor list")
    src = PM_EDITORS_JS.read_text(encoding="utf-8")
    block = src.split("window.PM_EDITORS", 1)[-1].split("];", 1)[0]
    seen, out = set(), []
    for line in block.splitlines():
        m = re.search(r'"([^"]+@[^"]+)"\s*,?\s*(?://\s*(.*))?$', line)
        if not m:
            continue
        email = m.group(1).strip().lower()
        if email in seen:
            continue
        seen.add(email)
        out.append([email, None, (m.group(2) or "").strip() or None])
    return out


# ── main ───────────────────────────────────────────────────────────────────
def main():
    # Only the workbooks that are actually present are loaded, and only their
    # tables are touched.
    #
    # This used to exit unless all three were found, which — once the path was
    # wrong and none were — was the only thing standing between a run and a
    # catastrophe: the truncate below names five tables unconditionally, so a
    # missing workbook would have emptied its table and inserted nothing back.
    # Skipping a table entirely is the safe reading of "that workbook isn't here".
    if not any((OKR_XLSX, DIR_XLSX, KPI_XLSX)):
        sys.exit(
            "No source workbooks found. Looked in:\n  "
            + "\n  ".join(str(d) for d in SEARCH_DIRS)
            + "\nExpected one or more of:\n"
              "  Profit.co Monthly Reports.xlsx\n"
              "  Student Services Org Directory.xlsx\n"
              "  Student Services KPIs.xlsx"
        )

    for label, path in (("OKRs", OKR_XLSX), ("Directory", DIR_XLSX), ("KPIs", KPI_XLSX)):
        print(f"  {label:10} {path if path else '— not found, its tables are left alone'}")
    print()

    okrs      = load_okrs()             if OKR_XLSX else None
    employees = load_employees()        if DIR_XLSX else None
    students  = load_student_employees() if DIR_XLSX else None
    org_chart = load_org_chart()        if DIR_XLSX else None
    kpis      = load_kpis()             if KPI_XLSX else None
    editors   = load_editors()

    # table -> (audit trigger name, rows) for everything we actually have data
    # for. Nothing else is truncated, disabled or re-enabled.
    loaded = [
        ("okrs",              "okrs_audit",              okrs),
        ("employees",         "employees_audit",         employees),
        ("student_employees", "student_employees_audit", students),
        ("org_chart_nodes",   "org_chart_nodes_audit",   org_chart),
        ("kpis",              "kpis_audit",              kpis),
    ]
    present = [(t, trg, rows) for t, trg, rows in loaded if rows is not None]

    parts = [
        "-- ═══════════ SEED DATA ═══════════\n"
        f"-- Generated {datetime.date.today().isoformat()} by supabase/import_sheets.py\n"
        "--\n"
        "-- DESTRUCTIVE. This truncates and reloads the tables listed below, which\n"
        "-- discards anything that lives only in the database — the Active/Inactive\n"
        "-- flags PMs set in the Hub included. It is a disaster-recovery tool for\n"
        "-- rebuilding from nothing, not the way to take on an updated workbook.\n"
        "-- The PM Hub is the source of truth for this data now. Running this\n"
        "-- against the live database would replace what PMs have maintained\n"
        "-- there with whatever a workbook happens to say.\n"
        "--\n"
        "-- Audit triggers are suspended so the import is not logged as user edits.\n\n"
        "begin;\n\n"
    ]
    parts += [f"alter table public.{t} disable trigger {trg};\n" for t, trg, _ in present]
    parts.append(
        "\ntruncate " + ",\n         ".join(f"public.{t}" for t, _, _ in present)
        + " restart identity;\n\n"
    )

    if okrs is not None:
        parts.append(insert("okrs", [
            "sort_order", "okr", "key_result", "sub_key_result", "sub_key_result_child",
            "period", "primary_stakeholder", "secondary_stakeholders", "project_manager",
            "type", "goal", "stretch_goal", "progress", "status", "trend", "comment",
            "update_date"], okrs))

    if employees is not None:
        parts.append(insert("employees", [
            "sort_order", "name", "role", "department", "employment_type",
            "primary_stakeholder", "sub_department", "contract_organization", "tier"], employees))

    if students is not None:
        parts.append(insert("student_employees", [
            "sort_order", "name", "job_name", "role_title", "sub_department",
            "supervisor", "department"], students))

    if org_chart is not None:
        parts.append(insert("org_chart_nodes", [
            "sort_order", "name", "role", "employee_status", "stewardships", "key_kpis",
            "reports_to", "department", "link", "key_responsibilities", "direct_reports"], org_chart))

    if kpis is not None:
        parts.append(insert("kpis", [
            "sort_order", "employee", "role", "department", "kpi_measure", "kpi_category",
            "category_type", "data_availability", "band_green", "band_yellow", "band_red",
            "tracking_status", "current_value", "data_source", "update_frequency",
            "update_date", "direction_hint", "green_cutoff", "red_cutoff"], kpis))

    if editors:
        parts.append(
            "insert into public.allowed_editors (email, full_name, note) values\n"
            + ",\n".join("  (" + ", ".join(sql(v) for v in e) + ")" for e in editors)
            + "\non conflict (email) do nothing;\n\n"
        )

    parts += [f"alter table public.{t} enable trigger {trg};\n" for t, trg, _ in present]
    parts.append("\ncommit;\n")

    # A partial run must not overwrite the committed full seed.
    #
    # seed.sql is the disaster-recovery artefact: every table, fully populated.
    # Running this with only one workbook present used to rewrite that file with
    # just the tables it had, silently discarding the OKR and KPI sections — the
    # database was never at risk (the truncate only names tables it is about to
    # fill) but the file was, and the file is the whole point of it.
    missing = [n for n, p in (("OKRs", OKR_XLSX), ("Directory", DIR_XLSX),
                              ("KPIs", KPI_XLSX)) if p is None]
    out = OUT
    if missing:
        out = OUT.with_name("seed-partial.sql")
        print(f"!! {', '.join(missing)} not found, so this is a PARTIAL seed.")
        print(f"!! Writing {out.name} instead of overwriting {OUT.name}.")
        print("!! It is safe to run — it only truncates the tables it refills —")
        print("!! but it is not a replacement for the full seed.\n")

    out.write_text("".join(parts), encoding="utf-8", newline="\n")

    print(f"wrote {out.relative_to(REPO)}")
    for table, _, rows in loaded:
        print(f"  {table:18}{len(rows) if rows is not None else '— skipped':>6}")
    print(f"  {'allowed_editors':18}{len(editors):>6}")
    if kpis:
        tracked = sum(1 for k in kpis if str(k[11]).strip().lower() == "tracking")
        print(f"    of which tracking: {tracked}")


if __name__ == "__main__":
    main()
