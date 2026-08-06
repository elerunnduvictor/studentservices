"""
Write the current directory state back into the Org Directory workbook.

The workbook had fallen behind the hub: reporting-line changes and two new
people existed only in directory/js/employees.js. This rewrites the Employee
Directory sheet so the workbook, the database and the hub all say the same
thing — after which the database is the source of truth and this is only
needed if someone edits the spreadsheet again.

Formatting, column widths and the surrounding blocks are left alone; only the
cell values in the staff block change.

    python supabase/update_workbook.py [--dry-run]
"""
import re
import shutil
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
WORKBOOK = REPO / "data-sources" / "Student Services Org Directory.xlsx"
EMPLOYEES_JS = REPO / "directory" / "js" / "employees.js"
SHEET = "Employee Directory"


def read_hub_employees():
    src = EMPLOYEES_JS.read_text(encoding="utf-8")
    pattern = re.compile(
        r'\{\s*name:\s*"([^"]*)",\s*role:\s*"([^"]*)",\s*dept:\s*"([^"]*)",\s*'
        r'type:\s*"([^"]*)",\s*stakeholder:\s*"([^"]*)",\s*subDept:\s*"([^"]*)",\s*'
        r'org:\s*"([^"]*)",\s*tier:\s*"([^"]*)"\s*\}'
    )
    return [
        dict(zip(("name", "role", "dept", "type", "stakeholder", "subDept", "org", "tier"),
                 (g.strip() for g in m.groups())))
        for m in pattern.finditer(src)
    ]


def find_header(ws, needle, max_scan=15):
    for ri in range(1, min(max_scan, ws.max_row) + 1):
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(ri, ci).value
            if v and str(v).strip().upper() == needle:
                return ri, ci
    raise SystemExit(f'Heading "{needle}" not found in {ws.title}')


def main():
    dry = "--dry-run" in sys.argv
    hub = read_hub_employees()

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[SHEET]
    hdr_row, name_col = find_header(ws, "NAME")
    stu_hdr, stu_col = find_header(ws, "EMPLOYEE NAME")

    students = set()
    for r in ws.iter_rows(min_row=stu_hdr + 1, min_col=stu_col, max_col=stu_col, values_only=True):
        if r[0]:
            students.add(str(r[0]).strip().lower())

    def is_student(p):
        n = p["name"].strip().lower()
        if p["type"].strip().lower() == "student employee":
            return True
        if n in students:
            return True
        short = " ".join(n.split()[:2])
        return any(s.startswith(short) for s in students)

    staff = [p for p in hub if not is_student(p)]

    # existing rows, by name
    existing = {}
    row = hdr_row + 1
    last_row = hdr_row
    while row <= ws.max_row:
        v = ws.cell(row, name_col).value
        if v and str(v).strip():
            existing[str(v).strip().lower()] = row
            last_row = row
        row += 1

    RENAMES = {"mandy schwab": "mandy poll schwab", "johanna relkin": "joanna relken"}
    for old, new in RENAMES.items():
        if old in existing and new not in existing:
            existing[new] = existing[old]

    # NAME, ROLE, DEPARTMENT, EMPLOYMENT TYPE, PRIMARY STAKEHOLDER,
    # SUB DEPT, CONTRACT ORGANIZATION, Tier Type
    ORDER = ["name", "role", "dept", "type", "stakeholder", "subDept", "org", "tier"]

    changed, added = 0, 0
    for p in staff:
        key = p["name"].strip().lower()
        r = existing.get(key)
        if r is None:
            last_row += 1
            r = last_row
            added += 1
            for i, f in enumerate(ORDER):
                ws.cell(r, name_col + i).value = p[f] or None
            print(f"  + row {r}: {p['name']} — {p['role']}")
            continue
        for i, f in enumerate(ORDER):
            cell = ws.cell(r, name_col + i)
            want = p[f] or None
            have = str(cell.value).strip() if cell.value is not None else None
            if (want or "") != (have or ""):
                if not dry:
                    cell.value = want
                changed += 1
                print(f"  ~ row {r} {p['name'][:26]:26} {ORDER[i]:12} {have!r} -> {want!r}")

    print(f"\n{changed} cell(s) updated, {added} row(s) added")

    if dry:
        print("(dry run — nothing written)")
        return

    backup = WORKBOOK.with_suffix(".xlsx.bak")
    shutil.copy2(WORKBOOK, backup)
    wb.save(WORKBOOK)
    print(f"saved {WORKBOOK.name}  (backup: {backup.name})")


if __name__ == "__main__":
    main()
